import os
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re

def orden_natural(texto):
    """ Función para ordenar cadenas con números de forma lógica (1, 2, 10...) """
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', texto)]

def cargar_configuracion():
    nombre_archivo = 'config.json'
    if not os.path.exists(nombre_archivo):
        print(f"Error: No se encontró el archivo '{nombre_archivo}'.")
        return None
    try:
        with open(nombre_archivo, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error al leer el JSON: {e}")
        return None


def formula_col_G(f):
    d = f"C{f}"
    tipos = [
        (f'OR(ISNUMBER(SEARCH("CEM IV ",{d})),ISNUMBER(SEARCH("CEM IV/",{d})))', '"CEM_IV"'),
        (f'OR(ISNUMBER(SEARCH("CEM I ",{d})),ISNUMBER(SEARCH("CEM I/",{d})),ISNUMBER(SEARCH("CEM III ",{d})),ISNUMBER(SEARCH("CEM III/",{d})))', '"CEM_I_III"'),
        (f'OR(ISNUMBER(SEARCH("CEM II ",{d})),ISNUMBER(SEARCH("CEM II/",{d})),ISNUMBER(SEARCH("CEM V ",{d})),ISNUMBER(SEARCH("CEM V/",{d})))', '"CEM_II_V"'),
        (f'OR(ISNUMBER(SEARCH("BL I ",{d})),ISNUMBER(SEARCH("BL I/",{d})))', '"BL_I"'),
        (f'OR(ISNUMBER(SEARCH("BL II ",{d})),ISNUMBER(SEARCH("BL II/",{d})))', '"BL_II"'),
        (f'ISNUMBER(SEARCH("BL 22,5 ",{d}))', '"BL_225"'),
        (f'ISNUMBER(SEARCH("Clinker blanco",{d}))', '"CLINKER_BLANCO"'),
        (f'OR(ISNUMBER(SEARCH("Clinker SR ",{d})),ISNUMBER(SEARCH("Clinker",{d})))', '"CLINKER"'),
        (f'OR(ISNUMBER(SEARCH("Puzolana MR",{d})),ISNUMBER(SEARCH("Puzolana SR",{d})))', '"PUZO_MR_SR"'),
        (f'ISNUMBER(SEARCH("Puzolana",{d}))', '"PUZO"'),
        (f'OR(ISNUMBER(SEARCH("Ceniza MR",{d})),ISNUMBER(SEARCH("Ceniza SR",{d})))', '"CENIZA_MR_SR"'),
        (f'AND(ISNUMBER(SEARCH("Ceniza",{d})),NOT(OR(ISNUMBER(SEARCH("MR",{d})),ISNUMBER(SEARCH("SR",{d})))))', '"CENIZA"'),
        (f'ISNUMBER(SEARCH("Escoria",{d}))', '"ESCORIA"'),
        (f'ISNUMBER(SEARCH("Caliza",{d}))', '"CALIZA"'),
        (f'OR(ISNUMBER(SEARCH("IV ",{d})),ISNUMBER(SEARCH("IV/",{d})))', '"CEM_IV"'),
        (f'OR(ISNUMBER(SEARCH("III ",{d})),ISNUMBER(SEARCH("III/",{d})))', '"CEM_I_III"'),
        (f'OR(ISNUMBER(SEARCH("II ",{d})),ISNUMBER(SEARCH("II/",{d})))', '"CEM_II_V"'),
        (f'OR(ISNUMBER(SEARCH("I ",{d})),ISNUMBER(SEARCH("I/",{d})))', '"CEM_I_III"'),
        (f'OR(ISNUMBER(SEARCH("V ",{d})),ISNUMBER(SEARCH("V/",{d})))', '"CEM_II_V"'),
    ]
    formula = '""'
    for cond, val in reversed(tipos):
        formula = f'IF({cond},{val},{formula})'
    return '=' + formula


def formula_col_H(f):
    d = f"C{f}"
    g = f"G{f}"
    marca_n = f'F{f}=$N$3'

    lh_i_iii = (f'AND(ISNUMBER(SEARCH("LH",{d})),'
                f'OR(ISNUMBER(SEARCH("CEM I ",{d})),ISNUMBER(SEARCH("CEM I/",{d})),'
                f'ISNUMBER(SEARCH("CEM III ",{d})),ISNUMBER(SEARCH("CEM III/",{d}))))')
    ba_i_iii = (f'AND(ISNUMBER(SEARCH(" (Ba)",{d})),'
                f'OR(ISNUMBER(SEARCH("CEM I ",{d})),ISNUMBER(SEARCH("CEM I/",{d})),'
                f'ISNUMBER(SEARCH("CEM III ",{d})),ISNUMBER(SEARCH("CEM III/",{d}))))')
    lh_ii_v  = (f'AND(ISNUMBER(SEARCH("LH",{d})),'
                f'OR(ISNUMBER(SEARCH("CEM II ",{d})),ISNUMBER(SEARCH("CEM II/",{d})),'
                f'ISNUMBER(SEARCH("CEM V ",{d})),ISNUMBER(SEARCH("CEM V/",{d}))))')

    cem_i_iii = (f'IF({lh_i_iii},IF({marca_n},16,15),'
                 f'IF({ba_i_iii},IF({marca_n},14,13),'
                 f'IF({marca_n},1,2)))')

    cem_ii_v = f'IF({lh_ii_v},IF({marca_n},18,17),IF({marca_n},3,4))'

    simples = [
        (f'{g}="CEM_IV"',         f'IF({marca_n},5,6)'),
        (f'{g}="BL_I"',           f'IF({marca_n},7,8)'),
        (f'{g}="BL_II"',          f'IF({marca_n},9,10)'),
        (f'{g}="BL_225"',         f'IF({marca_n},11,12)'),
        (f'{g}="CLINKER_BLANCO"', '22'),
        (f'{g}="CLINKER"',        '21'),
        (f'{g}="PUZO_MR_SR"',     '26'),
        (f'{g}="PUZO"',           '25'),
        (f'{g}="CENIZA_MR_SR"',   '24'),
        (f'{g}="CENIZA"',         '23'),
        (f'{g}="ESCORIA"',        '27'),
        (f'{g}="CALIZA"',         '28'),
    ]
    formula = '0'
    for cond, val in reversed(simples):
        formula = f'IF({cond},{val},{formula})'

    formula = f'IF({g}="CEM_II_V",{cem_ii_v},{formula})'
    formula = f'IF({g}="CEM_I_III",{cem_i_iii},{formula})'
    formula = f'IF(F{f}=$Q$3,29,{formula})'
    return '=' + formula


def formula_col_E(f):
    return f'=IF(H{f}=0,"",IF(F{f}=$O$3,$L$19,IF(F{f}=$P$3,$L$20,INDEX($L:$L,H{f}))))'


def _verificar_balance(nombre, formula):
    opens  = formula.count('(')
    closes = formula.count(')')
    if opens != closes:
        raise ValueError(f"Formula {nombre} desbalanceada: {opens} abre vs {closes} cierra")


def generar_informe():
    data_json = cargar_configuracion()
    if not data_json:
        return

    config = data_json.get("config")
    origen = config['carpeta_origen']
    salida = config['carpeta_salida']
    mes    = config['mes_proceso']
    anio   = config['anio_proceso']

    if not os.path.exists(salida):
        os.makedirs(salida)

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Informe Facturacion"

    tabla_p = data_json.get("tabla_precios", [])
    for i, item in enumerate(tabla_p, start=1):
        ws_out[f'K{i}'] = item['nombre']
        ws_out[f'L{i}'] = item['precio']

    tabla_r = data_json.get("tabla_referencias", [])
    for ref in tabla_r:
        ws_out[ref['celda']] = ref['valor']
        col = ref['celda'][0]
        ws_out[f'{col}2'] = ref['nombre']

    ws_out['G3'] = 'Tipo cemento'
    ws_out['G3'].font = Font(italic=True, color='808080')
    ws_out['H3'] = 'Fila $L'
    ws_out['H3'].font = Font(italic=True, color='808080')

    ws_out.merge_cells('A1:F1')
    titulo           = ws_out['A1']
    titulo.value     = f"Muestras Marca N (fecha toma {mes} {anio})"
    titulo.font      = Font(bold=True, size=14)
    titulo.fill      = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    titulo.alignment = Alignment(horizontal="center", vertical="center")

    fila_actual_salida = 5

    archivos = [f for f in os.listdir(origen) if f.lower().endswith(('.xlsx', '.xlsm'))]
    
    archivos.sort(key=orden_natural)

    for nombre_archivo in archivos:

        ruta_completa = os.path.join(origen, nombre_archivo)
        try:
            wb_in = openpyxl.load_workbook(ruta_completa, data_only=True)
            ws_in = wb_in.active

            fabrica = ws_in.cell(row=12, column=24).value or "Fabrica Sin Nombre"

            ws_out.merge_cells(f'A{fila_actual_salida}:F{fila_actual_salida}')
            celda_fab       = ws_out[f'A{fila_actual_salida}']
            celda_fab.value = fabrica
            celda_fab.font  = Font(bold=True, color="FFFFFF")
            celda_fab.fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            fila_actual_salida += 1

            inicio_bloque = fila_actual_salida
            hay_productos = False

            for r in range(53, ws_in.max_row + 1):
                codigo      = ws_in.cell(row=r, column=5).value
                nombre_prod = ws_in.cell(row=r, column=8).value
                val_g       = ws_in.cell(row=r, column=31).value
                val_as      = ws_in.cell(row=r, column=45).value

                if not codigo and not nombre_prod:
                    continue

                marca_final = ""
                if val_g is not None and str(val_g).strip() != "":
                    marca_final = 1
                elif val_as is not None and str(val_as).strip() != "":
                    rango_vacio = True
                    for c in range(27, 45):
                        if ws_in.cell(row=r, column=c).value not in (None, ""):
                            rango_vacio = False
                            break
                    if rango_vacio:
                        marca_final = 4

                f = fila_actual_salida
                hay_productos = True

                ws_out.merge_cells(f'A{f}:B{f}')
                ws_out[f'A{f}'] = codigo
                ws_out[f'C{f}'] = nombre_prod
                ws_out[f'F{f}'] = marca_final

                fg = formula_col_G(f)
                fh = formula_col_H(f)
                fe = formula_col_E(f)

                _verificar_balance(f'G{f}', fg)
                _verificar_balance(f'H{f}', fh)
                _verificar_balance(f'E{f}', fe)

                ws_out[f'G{f}'] = fg
                ws_out[f'H{f}'] = fh
                ws_out[f'E{f}'] = fe

                fila_actual_salida += 1

            if hay_productos:
                ws_out[f'A{fila_actual_salida}'] = "TOTAL"
                ws_out[f'A{fila_actual_salida}'].font = Font(bold=True)
                ws_out[f'E{fila_actual_salida}'] = f"=SUM(E{inicio_bloque}:E{fila_actual_salida-1})"
                ws_out[f'E{fila_actual_salida}'].font = Font(bold=True)
                fila_actual_salida += 2
            else:
                fila_actual_salida -= 1

        except Exception as e:
            print(f"Error procesando {nombre_archivo}: {e}")

    ws_out.column_dimensions['A'].width = 15
    ws_out.column_dimensions['C'].width = 50
    ws_out.column_dimensions['E'].width = 15
    ws_out.column_dimensions['G'].width = 18
    ws_out.column_dimensions['H'].width = 10
    ws_out.column_dimensions['K'].width = 25

    # Descomentar para ocultar las columnas auxiliares:
    ws_out.column_dimensions['G'].hidden = True
    ws_out.column_dimensions['H'].hidden = True

    nombre_final = f"Facturacion_{mes}_{anio}.xlsx"
    ruta_salida  = os.path.join(salida, nombre_final)
    wb_out.save(ruta_salida)
    print(f"Completado. Archivo generado en: {ruta_salida}")


if __name__ == "__main__":
    generar_informe()